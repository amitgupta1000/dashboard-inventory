"""
Data loader for market_data_hvb from Excel price reports
"""
import pandas as pd
import openpyxl
from datetime import datetime
from pathlib import Path
from sqlalchemy.orm import sessionmaker
import sys

# Add parent directory to path for imports
parent_dir = str(Path(__file__).parent.parent)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

from backend.database import get_engine, MarketDataHVB
from backend.ingestion_feedback import (
    validate_schema,
    create_ingestion_feedback,
    IngestionFeedback,
)


def extract_product_info(product_str: str) -> tuple[str, str]:
    """
    Extract product name and port from product string like 'TOLUENE - KANDLA'
    Returns (product_name, port)
    """
    if not product_str or not isinstance(product_str, str):
        return ('', '')
    
    # Split once on '-' and strip both parts to remove trailing/leading whitespace.
    cleaned = str(product_str).strip()
    if '-' not in cleaned:
        return (cleaned, '')

    product_name, port = cleaned.split('-', 1)
    product_name = product_name.strip()
    port = port.strip()
    
    return (product_name, port)


def load_market_data_from_excel(file_path: str, report_date: datetime.date = None) -> IngestionFeedback:
    """
    Load market data from Excel file into market_data_hvb table
    
    Args:
        file_path: Path to the Excel file
        report_date: Date for this report (defaults to today)
    
    Returns:
        IngestionFeedback with detailed feedback including schema validation,
        row counts, and error tracking.
    """
    try:
        if report_date is None:
            report_date = datetime.now().date()
        
        # Define expected columns (what we map from the Excel file)
        expected_columns = [
            'product', 'company', 'monthly_volumes', 'ready',
            'incoming_period_1', 'incoming_period_2', 'physical_stock',
            'pending_lifting', 'port_stock', 'replac_dollar', 'replace',
            'index', 'market_price', 'selling_p', 'current_month',
            'next_month', 'arrival_date'
        ]
        
        # Read Excel file with openpyxl to get exact cell values
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Extract USD/INR rate from last row, column M (13)
        last_row = ws.max_row
        usdinr_rate = ws.cell(last_row, 13).value  # Column M
        
        # Read data with pandas (excluding last row which is TOTAL)
        df = pd.read_excel(file_path, sheet_name=0)
        
        # Validate schema before processing
        actual_columns = [col for col in df.columns if not str(col).startswith('Unnamed')]
        schema_validation = validate_schema(
            expected_columns=expected_columns,
            actual_columns=actual_columns,
        )
        
        # Remove the last row (TOTAL)
        df = df.iloc[:-1]
        
        # Standardize expected column order and apply domain aliases.
        df.columns = [
            'product',               # 0
            'company',               # 1
            'monthly_volumes',       # 2
            'ready',                 # 3
            'incoming_period_1',     # 4 (alias of SECOND HALF MAY)
            'incoming_period_2',     # 5 (alias of JUNE 1ST HALF)
            'physical_stock',        # 6 (note: PHYSCIAL in original has typo)
            'pending_lifting',       # 7
            'port_stock',            # 8
            'replac_dollar',         # 9
            'replace',               # 10
            'index',                 # 11
            'market_price',          # 12
            'selling_p',             # 13
            'current_month',         # 14 (alias of MAY)
            'next_month',            # 15 (alias of JUNE)
            'arrival_date',          # 16
            'unnamed_17',            # 17 (ignore)
            'unnamed_18',            # 18 (ignore)
            'unnamed_19'             # 19 (ignore)
        ]
        
        # Extract product info
        df[['product_name', 'port']] = df['product'].apply(
            lambda x: pd.Series(extract_product_info(str(x)))
        )
        
        # Convert arrival_date to string, replacing NaN with None
        df['arrival_date'] = df['arrival_date'].apply(
            lambda x: str(x).strip() if pd.notna(x) and str(x).strip().lower() != 'nan' else None
        )
        
        # Add report metadata
        df['report_date'] = report_date
        df['usdinr_rate'] = usdinr_rate
        
        # Get database engine and create session
        engine = get_engine()
        
        # Create table if it doesn't exist
        MarketDataHVB.__table__.create(engine, checkfirst=True)
        
        Session = sessionmaker(bind=engine)
        session = Session()
        
        inserted = 0
        failed = 0
        errors = []
        
        try:
            # Load data
            for idx, row in df.iterrows():
                try:
                    market_data = MarketDataHVB(
                        report_date=report_date,
                        product=row['product'],
                        company=row['company'],
                        monthly_volumes=_to_float(row['monthly_volumes']),
                        ready=_to_float(row['ready']),
                        incoming_period_1=_to_float(row['incoming_period_1']),
                        incoming_period_2=_to_float(row['incoming_period_2']),
                        physical_stock=_to_float(row['physical_stock']),
                        pending_lifting=_to_float(row['pending_lifting']),
                        port_stock=_to_float(row['port_stock']),
                        replac_dollar=_to_float(row['replac_dollar']),
                        replace=_to_float(row['replace']),
                        index=_to_float(row['index']),
                        market_price=_to_float(row['market_price']),
                        selling_p=_to_float(row['selling_p']),
                        current_month=_to_float(row['current_month']),
                        next_month=_to_float(row['next_month']),
                        arrival_date=row['arrival_date'],
                        product_name=row['product_name'],
                        port=row['port'],
                        usdinr_rate=_to_float(usdinr_rate),
                    )
                    session.add(market_data)
                    inserted += 1
                except Exception as e:
                    failed += 1
                    errors.append(f"Row {idx}: {str(e)}")
            
            session.commit()
            
            # Determine status
            total_rows = inserted + failed
            if failed == 0:
                status = "success"
                message = f"✅ Successfully loaded {inserted} market data records"
            elif inserted > 0:
                status = "partial_success"
                message = f"⚠️ Partially loaded: {inserted}/{total_rows} rows succeeded"
            else:
                status = "failed"
                message = f"❌ Failed to load market data: all {total_rows} rows failed"
            
            return create_ingestion_feedback(
                status=status,
                message=message,
                total_rows=total_rows,
                inserted=inserted,
                updated=0,
                failed=failed,
                schema_validation=schema_validation,
                errors=errors,
                report_date=report_date.isoformat(),
                source_file=Path(file_path).name,
                destination_table="market_data_hvb",
            )
        
        except Exception as e:
            session.rollback()
            return create_ingestion_feedback(
                status="failed",
                message=f"❌ Database error: {str(e)}",
                total_rows=len(df),
                errors=[str(e)],
                schema_validation=schema_validation,
                report_date=report_date.isoformat(),
                source_file=Path(file_path).name,
                destination_table="market_data_hvb",
            )
        finally:
            session.close()
    
    except Exception as e:
        return create_ingestion_feedback(
            status="failed",
            message=f"❌ Failed to read Excel file: {str(e)}",
            total_rows=0,
            errors=[str(e)],
            source_file=Path(file_path).name,
            destination_table="market_data_hvb",
        )


def _to_float(value):
    """Convert value to float, handling None and string values"""
    if value is None:
        return None
    if isinstance(value, float):
        if pd.isna(value):
            return None
        return value
    if isinstance(value, int):
        return float(value)
    if isinstance(value, str):
        s = str(value).strip()
        if s.lower() in ['nan', '', 'none']:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


if __name__ == '__main__':
    # Example usage
    file_path = 'data_files/DAILY PRICE REPORT 22-05-2026.xlsx'
    
    if not Path(file_path).exists():
        print(f"❌ File not found: {file_path}")
        sys.exit(1)
    
    # Extract date from filename (22-05-2026)
    filename = Path(file_path).stem
    try:
        date_str = filename.split()[-1]  # Get '22-05-2026'
        day, month, year = date_str.split('-')
        report_date = datetime(int(year), int(month), int(day)).date()
    except:
        report_date = None
    
    load_market_data_from_excel(file_path, report_date)
