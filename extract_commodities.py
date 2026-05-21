import openpyxl
from collections import defaultdict
import csv

file_path = '12-5-26.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['12-05-2026']

# Extract unique commodities
commodities = set()
commodity_details = defaultdict(lambda: {'terminals': set(), 'ports': set(), 'count': 0})

for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
    if row[3]:  # Column D - PRODUCT NAME
        commodity_name = str(row[3]).strip()
        terminal = str(row[10]).strip() if row[10] else 'Unknown'
        port = str(row[4]).strip() if row[4] else 'Unknown'
        
        commodities.add(commodity_name)
        commodity_details[commodity_name]['terminals'].add(terminal)
        commodity_details[commodity_name]['ports'].add(port)
        commodity_details[commodity_name]['count'] += 1

# Output to CSV
output_csv_path = 'unique_commodities.csv'
with open(output_csv_path, 'w', newline='', encoding='utf-8') as csvfile:
    csv_writer = csv.writer(csvfile)
    csv_writer.writerow(['Commodity Name'])  # Write header
    for commodity in sorted(commodities):
        csv_writer.writerow([commodity])

print(f'🏷️  Unique Commodities Found: {len(commodities)}')
print(f'CSV file created at: {output_csv_path}')
print("Contents:")
with open(output_csv_path, 'r', encoding='utf-8') as f:
    print(f.read())
